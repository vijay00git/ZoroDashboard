#!/usr/bin/env python3
"""Automation Finder — Google Sheet scan.

Downloads the shared "Automation Coverage" Google Sheet as .xlsx and reads
every sprint/regression tab into one flat, case-ID-keyed JSON blob. Kept as a
standalone script (rather than a Node dependency) because openpyxl is already
the best tool on hand for walking a multi-tab workbook, and this only ever
needs to run as a child process from server.js's sync job.

Usage: python3 automation_sheet_scan.py <sheet_url> <output_json_path>

On success, writes the JSON blob to <output_json_path> and exits 0.
On failure, writes {"error": "..."} to <output_json_path> and exits 1 — the
caller should always read the output file, not stdout, to learn what happened.
"""
import json
import re
import sys
import urllib.request
from io import BytesIO

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Tabs that hold documentation/legend content rather than case rows.
META_TABS = {"guideline", "status", "tempids"}

# Column layout shared by every data tab (1-indexed, matches the sheet's own
# lettering A..M) — confirmed against the live sheet before building this.
COL = {
    "module": 1,
    "case_id": 2,
    "case_status": 3,
    "cloud_status": 4,
    "cloud_automated_in": 5,
    "cloud_remarks": 6,
    "cloud_qa": 7,
    "cloud_device_status": 8,
    "device_case_id": 9,
    "device_status": 10,
    "device_automated_in": 11,
    "device_qa": 12,
    "device_remarks": 13,
}

def header_matches_layout(ws):
    """Positional-but-fuzzy check: does this tab's row 1 look like the
    standard Cloud/Device layout at the columns COL expects? Different tabs
    phrase headers slightly differently ("Test Rail ID" vs "Test Case Id"),
    so this matches on a keyword per column rather than an exact string —
    but still requires the keyword at the *specific* column, so a
    genuinely different layout (e.g. one with no Cloud section at all, or
    columns shifted) is correctly rejected rather than mis-mapped."""
    def cell(col):
        return str(ws.cell(row=1, column=col).value or "").lower()
    return (
        "id" in cell(COL["case_id"])
        and "automation" in cell(COL["cloud_status"])
        and "qa" in cell(COL["cloud_qa"])
        and "qa" in cell(COL["device_qa"])
    )


CASE_ID_RE = re.compile(r"^C\s*\d{4,}$", re.IGNORECASE)
DATE_RE = re.compile(r"\d{1,2}[-/]\d{1,2}[-/]\d{2,4}")
STATUS_WORDS = {"done", "pending", "blocked", "na", "n/a", "not applicable", "tbd", "sanity"}


def qa_issue(value):
    """Returns a short reason string if `value` doesn't look like a person's
    name typed into a QA column, else None."""
    if not value:
        return None
    v = str(value).strip()
    if not v:
        return None
    if CASE_ID_RE.match(v):
        return "looks like a case ID, not a name"
    if v.lower() in STATUS_WORDS:
        return "looks like a status word, not a name"
    if DATE_RE.search(v):
        return "contains a date, not just a name"
    return None


def normalize_case_id(raw):
    if raw is None:
        return None
    v = str(raw).strip()
    if not v:
        return None
    if not v.upper().startswith("C"):
        v = "C" + v
    return v.upper()


def fetch_workbook(url):
    export_url = url
    if "export?format=xlsx" not in url:
        m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
        if not m:
            raise ValueError(f"Could not find a spreadsheet ID in URL: {url}")
        export_url = f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"
    with urllib.request.urlopen(export_url, timeout=60) as resp:
        data = resp.read()
    return openpyxl.load_workbook(BytesIO(data), data_only=True)


def scan(url):
    if openpyxl is None:
        raise RuntimeError("openpyxl is not installed (pip install openpyxl)")

    wb = fetch_workbook(url)

    cases = {}
    scanned_tabs = []
    skipped_tabs = []

    for name in wb.sheetnames:
        if name.strip().lower() in META_TABS:
            skipped_tabs.append({"tab": name, "reason": "meta tab"})
            continue
        ws = wb[name]
        if not header_matches_layout(ws):
            skipped_tabs.append({"tab": name, "reason": "unexpected header layout"})
            continue
        scanned_tabs.append(name)

        for row in range(2, ws.max_row + 1):
            case_id = normalize_case_id(ws.cell(row=row, column=COL["case_id"]).value)
            if not case_id:
                continue

            cloud_qa = ws.cell(row=row, column=COL["cloud_qa"]).value
            device_qa = ws.cell(row=row, column=COL["device_qa"]).value
            # Column C — feasibility ("Automatable" / "Not Automatable" /
            # "Blocked") — distinct from column D, which is *progress*
            # ("Done" / "Pending" / ...). A "Not Automatable" case must never
            # be surfaced as a pick-able "not automated yet" candidate.
            record = {
                "module": ws.cell(row=row, column=COL["module"]).value,
                "caseStatus": ws.cell(row=row, column=COL["case_status"]).value,
                "cloudStatus": ws.cell(row=row, column=COL["cloud_status"]).value,
                "cloudAutomatedIn": ws.cell(row=row, column=COL["cloud_automated_in"]).value,
                "cloudRemarks": ws.cell(row=row, column=COL["cloud_remarks"]).value,
                "cloudQA": cloud_qa,
                "cloudDeviceStatus": ws.cell(row=row, column=COL["cloud_device_status"]).value,
                "deviceCaseId": ws.cell(row=row, column=COL["device_case_id"]).value,
                "deviceStatus": ws.cell(row=row, column=COL["device_status"]).value,
                "deviceAutomatedIn": ws.cell(row=row, column=COL["device_automated_in"]).value,
                "deviceQA": device_qa,
                "deviceRemarks": ws.cell(row=row, column=COL["device_remarks"]).value,
                "sheetTab": name,
                "rowNum": row,
            }
            cloud_issue = qa_issue(cloud_qa)
            device_issue = qa_issue(device_qa)
            qa_data_issues = []
            if cloud_issue:
                qa_data_issues.append(f"Cloud QA: {cloud_issue} ({cloud_qa!r})")
            if device_issue:
                qa_data_issues.append(f"Device QA: {device_issue} ({device_qa!r})")
            record["qaDataIssues"] = qa_data_issues

            if case_id in cases:
                cases[case_id]["duplicate"] = True
                cases[case_id]["duplicateRows"].append({"sheetTab": name, "rowNum": row})
            else:
                record["duplicate"] = False
                record["duplicateRows"] = []
                cases[case_id] = record

    return {
        "generatedAt": __import__("datetime").datetime.utcnow().isoformat() + "Z",
        "sheetUrl": url,
        "scannedTabs": scanned_tabs,
        "skippedTabs": skipped_tabs,
        "caseCount": len(cases),
        "cases": cases,
    }


def main():
    if len(sys.argv) != 3:
        print("Usage: automation_sheet_scan.py <sheet_url> <output_json_path>", file=sys.stderr)
        sys.exit(2)
    url, out_path = sys.argv[1], sys.argv[2]
    try:
        result = scan(url)
        with open(out_path, "w") as f:
            json.dump(result, f)
    except Exception as exc:  # noqa: BLE001 — this process's only job is to report success/failure to its caller
        with open(out_path, "w") as f:
            json.dump({"error": str(exc)}, f)
        sys.exit(1)


if __name__ == "__main__":
    main()
